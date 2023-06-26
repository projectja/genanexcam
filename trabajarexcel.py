# -*- coding: utf-8 -*-



"""
Spyder Editor

This is a temporary script file
"""
from docxtpl import DocxTemplate
import openpyxl as ox
import os
import pandas as pd
import pdfquery
import win32com.client
from jinja2 import Environment, FileSystemLoader
#pip install docx2pdf
from docx2pdf import convert
import base64
import glob
# move files
import shutil





cif = ""
cifglobal = ""
# esta es la ruta que se utiliza para el cif que estamos tratando
global rutaendiscoglobal 
fecha_documento = ''
max_score = 100
test_name =  "Python Challengue"
beneficiario = "Alfonso Raggio"
students = [
    {"name":"alfonso", "score": 90},
    {"name":"Manuel", "score": 50},
    {"name":"oscar", "score": 88}
]
environment = Environment(loader=FileSystemLoader(""))
print(" environment " + str(environment))
results_template = environment.get_template("template_1.htm")
results_filename = "anexo19.htm"
context = {
    "student": beneficiario,
    "test_name": test_name,
    "max_score": max_score,
}
with open(results_filename, mode="w", encoding="utf-8") as results:
    results.write(results_template.render(context))
    print(f"... wrote {results_filename}")




outlook = win32com.client.Dispatch('Outlook.Application')

doc = DocxTemplate('Anexo 03. Ficha de Empresa 2022-v2.docx')
CF = pd.DataFrame()



# datosDeExcel = pd.read_excel('example.xlsx')
# workBook = openpyxl.load_workbook('example.xlsx')
# workSheet = workBook.active
# workSheet1 = workBook['Sheet1']
# cell_value = workSheet1.cell(2,1) 

# print ('el tipo es ')
# print(workSheet1['B1'].value)
# workSheet2 = workBook['Sheet2']
# workSheet2['B2'].value = 2
# print(workSheet2['B2'].value)
# print('hola')


os.chdir('C:\\Users\\alfonso.DESKTOP-0B71OCJ\\OneDrive\\Desarrollo\\python\\automatizacion\\generar-anexos')
path = 'C:\\Users\\alfonso.DESKTOP-0B71OCJ\\OneDrive\\Desarrollo\\python\\automatizacion\\generar-anexos'

# plantillas
doc = DocxTemplate('my_word_template.docx')


# almacenará el numero de fila en la pestaña alta_usuario   
global numero_fila
numero_fila = 0


# CREAMOS DATAFRAMES de EMPRESAS Y USUARIOS ASOCIADOS

data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME = pd.read_excel('LISTA-ASESORE-GENERACION-DOCUMENTOS-2022-BAK-ANTES-DE-ANEXO-20-Y-21-AÑO-2023-V3.xlsx', sheet_name='LISTA-ASESORE-GENERACION-DOCUME', usecols = 'A:DY',header = 0)
# data_oap_tab_alta_usuario = pd.read_excel('10044_Huelva_Registro_Actividades_OAs_202112_v2.5.xlsx', sheet_name='Alta_Usuario', usecols = 'E:N',header = 3)
#data_oap_tab_actividades_individuales = pd.read_excel('10044_Huelva_Registro_Actividades_OAs_202112_v2.5.xlsx', sheet_name='Actividades individuales', usecols = 'E:T',header = 3)




# ==========================
# OBENEMOS EL MÁXIMO NUMERO DE FILAS
# numero de filas:
# row_count = data_oap_tab_alta_usuario.shape[0]
# al sumarles 5, le daremos la posición exacta en la que tiene que crear una fila en la pestaña de usuarios
# si el NIF no existiera

# row_count = row_count + 5
# print ("nuero de filas para SABER CUAL ES LA ULTIMA " + str(row_count))
# ==========================


# ================================= actualizar excel ======================
# =========================================================================
# =========================================================================
""" 
def update_spreadsheet(path, _df=CF, startcol:int=1, startrow:int=1, sheet_name:str ="Alta_Usuario", numFila:int=1):
    file =  path + '\\10044_Huelva_Registro_Actividades_OAs_202112_v2.5.xlsx'
    print("====================== UPDATE_SPREADSHEET =====================")
    
    #========================================================
    print (" file     " +  file)    
    # ESTO LO PUSE PARA CONTROLAR QUE EL ARCHIVO QUE SE 
    # ACCEDE SEA UN ARCHIVO O NO PORQUE DARÍA ERROR, DE ESTA FORMA LO COMPROBAMOS
    # PERO REALMENTE ESTO SE HIZO PARA DEBUGEAR
    # PARA METERLO EN PRODD HABRÍA QWUE TRATARLO
    assert os.path.isfile(file)
    with open(file, "r") as f:
        pass        
    #===================================================================
    wb = ox.load_workbook(file)    
    
    # SOLO DEBUG - print ("ruta completa " + str(path + '\\10044_Huelva_Registro_Actividades_OAs_202112_v2.5.xlsx'))
    print (" NUMERO DE FILA  EN UPDATE SPREADSHEET  " + str(numFila))
    ws=wb[sheet_name]
    #====================================================================
    
    
    print (" valor de un DATAFRAME " + str(_df)) 
    
    
    # RECORRER EL EXCCEL Y EL DATAFRAME
    for row in range(0, _df.shape[0]): #For each row in the dataframe
        for col in range(0, _df.shape[1]): #For each column in the dataframe
            print( " ===== DENTRO DEL FOR DE RELLENO DE EXCEL =======")
            # ws.cell
            # ws.cell(row = startrow + row, column = startcol + col).value = _df.iat[row, col]
            # _df.iat[0, 2]  => es la columna 2 que es el DNI (porque empieza a contar de iazquierda a derecha)
            #  ws.cell almacenará los valores en el excel,
            ws.cell(150, 5).value = _df.iat[0, 2] # NIF POSICION 3 EN EXCEL equivalente a 2 en una lista
            # FECHA, LA DEJAMOS EN BLANCO   ws.cell(150, 6).value = _df.iat[0, 3]
            ws.cell(150, 7).value = _df.iat[0, 4]  # tratamiento
            ws.cell(150, 8).value = _df.iat[0, 5]  # nombre
            # 8  - primer apellido
            # 9 - sedgundo a pellido
            ws.cell(150, 11).value = _df.iat[0, 2] # nif
            ws.cell(150, 12).value = _df.iat[0, 6]  #cargo
            ws.cell(150, 13).value = _df.iat[0, 7]  #coreo electronico
            
            
            # estructura del _df
            #[razons,estadonif,nifempresa,fecha,tratamiento,nombre,primerapellido,segundoapellido,nif,cargo,correoelectronico]],columns=['Razón Social','Estado NIF','documento_soclitiante','Fecha (DD/MM/AA)','Tratamiento','Nombre','Primer Apellido','Segundo Apellido','NIF','Cargo','Correo electrónico'])    
            
            # estructura del df equivalente a ALTA_USUARIO 
            #df_tab_AltaUsuario_fila_modificada_o_nueva =
            #pd.DataFrame([[razons - ro ,estadonif - ro ,nifempresa (columna 5 EN EXCEL),fecha,tratamiento,nombre,
            #primerapellido,      #segundoapellido,nif,cargo,correoelectronico]],
            #columns=['Razón Social','Estado NIF','documento_soclitiante','Fecha (DD/MM/AA)','Tratamiento','Nombre',
            #         'Primer Apellido','Segundo Apellido','NIF','Cargo','Correo electrónico'])    
            
            # print("valor de la celda " + str(ws.cell(row = startrow + row, column = startcol + col).value))
            # print("valor de la celda " + str(_df.iat[0, 3]) + "excel " + str(ws.cell(150,5).value))
            print(col)
            print(_df.iat[0,col])
    wb.save(file)

 """





# =================================== leer excel ===========
def read_data_pandas():
    data = pd.read_excel('example.xlsx')
    return data


def read_data_pandas_oap():
        # extraer datos a variable data1 de la oap
        # en la siguiente línea header, indica en qué linea empiezan los indices.
      
        print(data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME)
        return data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME, # data_oap_tab_alta_usuario, data_oap_tab_actividades_individuales

    
def modify_data_pandas(cif_a_buscar,data_oap_tab_alta_usuario):    
    print("======================MODIFY_DATA_PANDAS =====================")
    # Creamos los tres dataframes, uno por cada pestaña.

    # data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME, df_oap_alta_usuario, df_oap_ai # = read_data_pandas_oap()    
    #data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME
    #data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME 
    # = read_data_pandas_oap()    
    
    # ========= Si lo encontramos, creamos un dataframe con esa fila para la pestaña EMPRESA y USUARIO  
    # == Cargamos en DF datos TAB EMPRESA
   ######### df_tab_empresa_fila_buscada = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.loc[data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME['documento_soclitiante'] == cif_a_buscar]   

    
    # INICIALIZAMOS aqui VARIABLES para crear un dataframe vacío con su cabecera
    razons = ''
    estadonif = ''
    nifempresa = ''
    fecha = ''
    tratamiento = ''
    nombre = ''        
    primerapellido = ''
    segundoapellido = ''
    nif = ''
    cargo = ''
    correoelectronico = ''
           
        
    
    # ========================================   
    # ========================================
    # ========================================   CREAR EL DATAFRAME CON ESTRUCTURA TAB ALTA_USUARIO VACIO
    
    df_tab_AltaUsuario_fila_modificada_o_nueva = pd.DataFrame([[razons,estadonif,nifempresa,fecha,tratamiento,nombre,primerapellido,segundoapellido,nif,cargo,correoelectronico]],columns=['Razón Social','Estado NIF','documento_soclitiante','Fecha (DD/MM/AA)','Tratamiento','Nombre','Primer Apellido','Segundo Apellido','NIF','Cargo','Correo electrónico'])    
    
    
    # ======= obtenemos numero de fila en TAB EMPRESA para ese CIF
    # obtenemos el numero el índice que nos dará el NUMERO DE FILA en el excel
    # Esto es importante para poder usarlo después en el excel. Quedará almacenado en numero_fila
    """ for row in df_tab_empresa_fila_buscada.index:
         numero_fila_dfAltaEmpresa = row
         print(" fila  alta empresa indice " + str(row), end = " ") """

    # buscamos el CIF buscado en la lista de empresas para ver si ya existía  
   
    # si encuentra el CIF en el dataframe de empresas (ya se dio de alta), entre en el IF
    print("==  CIF a buscar es        " + cif_a_buscar)
    # imprimir columnas específicas de un dataframe
    #print(data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME[['documento_soclitiante','Nombre o razón social']])
     
    # ============== CIF existe en empresa y no en usuarios => lo daremos alta         
    if cif_a_buscar in data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.values and cif_a_buscar not in data_oap_tab_alta_usuario.values:
        print("==== DENTRO DEL IF CIF_a_BUSCAR IN DF_OAP_ALTA_USUARIO ===")
        print("== CIF NO EXISTE en TAB USUARIOS " + cif_a_buscar)
        # aqui crearemos la ficha en word para un nuevo asesoramiento                   
              
        # YA sabemos que el NIF está en LISTA-ASESORE-GENERACION-DOCUME
        # aqui busca el mismo NIF pero en la pestaña USUARIOS, PUEDE SER QUE existiera o que no existiera porque
        # no tenga ningun asesoramiento todavia y lo estemos dando de alta.
        # aqui es donde se da de alta el mismo NIF en la pestaña empresa. En el caso de autonomos viene bien porque
        # son los mismos datos
        # si no existe se rellenará al final,
        
                    
                  
        # ========= NOS DA EL NÚMERO DE FILA del CIF en TAB ALTA USUARIO    
        # obtenemos el numero el índice que nos dará el NUMERO DE FILA en el excel
        # Esto es importante para poder usarlo después en el excel. Quedará almacenado en numero_fila
        for row in df_tab_AltaUsuario_fila_modificada_o_nueva.index:
            # continene que FILA en la que se encuentra el NIF encontrado en el excel, pestaña ALTA_USUARIO
            # para sai poder actualizarlo
            numero_fila_dfAltaUsuario = row
            print(" fila alta usuario  indice " + str(row), end = " ")      
            
         
        # == Cargamos en DF datos TAB EMPLEADOS        
        """  df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario,'documento_soclitiante'] = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.at[numero_fila_dfAltaEmpresa,'documento_soclitiante']
        df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario,'Razón social'] = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.at[numero_fila_dfAltaEmpresa,'Nombre o razón social']
        df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario,'Tratamiento'] = 'Sr.'
        df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario,'Nombre'] = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.at[numero_fila_dfAltaEmpresa,'Nombre o razón social']
         """
                
                 
    else:
        
        # ========================= EXISTE EN TAB LISTA-ASESORE-GENERACION-DOCUME PERO TAMBIÉN EN TAB ALTA_USUARIO
        #===========  Significa que el usuario ya ha tenido un asesoramiento, se exportará a word la ficha
        if cif_a_buscar in data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.values and cif_a_buscar in data_oap_tab_alta_usuario.values:        
            print("====== DENTRO DEL ELSE IF CIF_a_BUSCAR IN DF_OAP_ALTA_USUARIO ===")
            print("=== El usuario ya existe en ALTA_USUARIO =====")
            print("===  aqui lo que hacemos es sacar el word - es una nueva visita a la OAP ==")
            
            # aquí asignamos los datos que ya existen en la FICHA DE EMPRESA a la FICHA DE USAURIO
            numero_fila_dfAltaUsuario = '0'
            # df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario, 'documento_soclitiante'] = df_tab_empresa_fila_buscada.at[numero_fila_dfAltaEmpresa,'documento_soclitiante'] 
            # print("hola")
            print(df_tab_AltaUsuario_fila_modificada_o_nueva)
            # df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario, 'documento_soclitiante'] = df_tab_empresa_fila_buscada.at[numero_fila_dfAltaEmpresa,'documento_soclitiante'] 
            # df_tab_AltaUsuario_fila_modificada_o_nueva.at[numero_fila_dfAltaUsuario, 'documento_soclitiante'] = df_tab_empresa_fila_buscada.at[numero_fila_dfAltaEmpresa,'documento_soclitiante'] 
            
            
            print('==========================')
            # print(df_tab_empresa_fila_buscada)
            print(df_tab_AltaUsuario_fila_modificada_o_nueva)
          
            # AQUI ESTÁN TODOS LOS USUARIOS DE LA PESTAÑA        
            print ( "DATAFRAME TAB ALTA DE USAURIO " + str(df_oap_alta_usuario))
            
            # AQUI ESTA LA FILA QUE COINCIDE CON EL CIF
            print ( "INDICE NIF BUSCADO " + str(df_tab_AltaUsuario_fila_modificada_o_nueva))
            # campo_numero_empleados = str(df_tab_empresa_fila_buscada.iloc[0]['Empleados empresa'])                 
            # campo_nif_usuario_asesorado = str(df_tab_AltaUsuario_fila_modificada_o_nueva.iloc[0]['NIF'] )       
            
            #print("Empresa asesorada - NUMERO DE EMPLEADOS " + campo_numero_empleados )    
            #print("Persona que recibe el asesoramiento - DNI " + str(campo_nif_usuario_asesorado))         
            
       #else:               
            
            
            #d AQUI ES DONDE DA ERROR ============ > 
            #df_tab_AltaUsuario_fila_modificada_o_nueva.at['0','documento_soclitiante']= data_oap_tab_alta_usuario.at['0','documento_soclitiante']
            #df_tab_AltaUsuario_fila_modificada_o_nueva.at['0', 'documento_soclitiante'] = df_tab_empresa_fila_buscada.at[numero_fila_dfAltaEmpresa,'documento_soclitiante'] 
                        
        # df_tab_AltaUsuario_fila_modificada_o_nueva = data_oap_tab_alta_usuario.at['0','documento_soclitiante'] = 111
            print ("ESTOY EN EL ELSE " + str(df_tab_AltaUsuario_fila_modificada_o_nueva))
            
            #numero_fila_dfAltaUsuario = row_count
            
             
        
    # VAMOS A PREPARAR LOS DATOS DE LA PESTAÑA ALTA_USUARIO. EN PRINCIPIO CONSIDERAMOS QUE LOS DATOS
    # DE LA EMPRESA DE LA PESTAÑA EMPRESA SERÁN LOS MISMOS DATOS TAMBIÉN DEL QUE CONSULTA, O SEA, QUE ES AUTONOMO    
   
    print("==== terminado el IF/ELSE ==")
    
    # ===============================
    # ===============================
    # ===============================
    # ===============================  ACTUALIZAR EL EXCEL
    # ===============================
    # ===============================
    # ===============================
   
    #update_spreadsheet(path,df_tab_AltaUsuario_fila_modificada_o_nueva ,0,1,"Alta_Usuario",numero_fila_dfAltaUsuario)            
    return data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME

    # FIN FUNCION =================  FIN FUNCION  modify_data_pandas

def crear_draftcorreo_19():
    send_account = None
    ruta = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/generated_anexo19_" + cifglobal + ".docx"
    convert(ruta, "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/anexo19.pdf")
    for account in outlook.Session.Accounts:
        if account.DisplayName == 'alfonso.raggio@camarahuelva.com':
            send_account = account
            break

    mail_item = outlook.CreateItem(0)   # 0: olMailItem

    # mail_item.SendUsingAccount = send_account not working
    # the following statement performs the function instead
    mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))

    mail_item.Recipients.Add('alfonso.raggio@camarahuelva.com')
    mail_item.Subject = 'Test sending using particular account'
    mail_item.BodyFormat = 2   # 2: Html format
    attachment = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/anexo19.pdf"
    attachment2 = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/ppi.pdf"
    mail_item.Attachments.Add(attachment)
    mail_item.Attachments.Add(attachment2)
    filename = "innocamaras.png"                   
    attachment3 =  'C:\\Users\\alfonso.DESKTOP-0B71OCJ\\OneDrive\\Desarrollo\\python\\automatizacion\\generar-anexos\\img\\innocamaras.png'
    

    attach = mail_item.Attachments.Add(attachment3)  
    attach.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "MyId1")

    # data:image/png;base64,

    with open(attachment3, "rb") as image:
            image_string = base64.b64encode(image.read())
            # imagede_string = base64.b64decode(image_string))

    


    mail_item.HTMLBody ='''
     <html><body>
	<p></p>
	<p>


</p><p>Buenos
días.</p>

<p>Adjunto
les remito la documentación a firmar para que puedan comenzar la implantación
de los proyectos que nos indicó.</p>

<p>Ruego comprueben bien los datos antes de su firma ya que en
ella se estipulan los proyectos y cuantías económicas autorizadas por esta
Cámara. <b><u>Una vez firmado por ambas partes, podrán comenzar a ejecutar
gastos. Ningún gasto realizado con anterioridad a la firma de los documentos
por ambas partes, no será admitido.</u></b></p>

<p>Los
documentos&nbsp; en formato PDF pueden remitirlos firmados con <b><u>su
certificado digital. </u></b></p>

<p>A continuación recibirá un email indicando las obligaciones de publicidad de
la UE que deben cumplir.</p>

<p>Les recordamos que los plazas de ejecución y justificación son los
siguientes.</p>

<table style="background-color:#FFFFE0;" border="1">
 <tbody><tr>
  <td valign="top">
  <p><b>FECHA FIN DE
  EJECUCIÓN Y PAGOS</b></p>
  </td>
  <td valign="top">
  <p><b>31 de agosto
  de 2023</b></p>
  </td>
 </tr>
 <tr>
  <td valign="top">
  <p><b>PLAZO LÍMITE
  DE JUSTIFICACIÓN</b></p>
  </td>
  <td valign="top">
  <p><b>15 de
  septiembre de 2023</b></p>
  </td>
 </tr>
</tbody></table>

<p>Le adjuntamos también el documento donde se detalla algunas consideraciones
a tener en cuenta para el pago y justificación de las facturas. Recomendamos su
lectura para evitar incidencias en la justificación.</p>

<p>Saludos y gracias de antemano. </p>
''' +  '<p>' +   '</p>  <img src="data:image/png;base64, ' + '' + str(image_string.decode('utf-8')) + '">'


# <p></p><p></p><p>&nbsp;</p> </body></html>    
    mail_item.Save()
    return


def crear_draft_anexo18():

   adjuntos = []
   send_account = None
   global rutaendiscoglobal 
   listddp = glob.glob(rutaendiscoglobal + '/' + '*ddp*.docx')
   listdiag =  glob.glob(rutaendiscoglobal + '/' + '*diagnostico*.docx')
   #print("listado de ddp " + listddp)
   #print("Diagnostico " + listdiag[0])
   print("ruta de los anexo 18 " + rutaendiscoglobal)
   wait = input("Press Enter to continue." )
   print("something")
   # shutil.move(rutaorigen + listdiag[0], rutaendiscoglobal + '/' + listdiag[0])
   for file in listddp:
       adjuntos.append(file)    
  
   for account in outlook.Session.Accounts:
        if account.DisplayName == 'alfonso.raggio@camarahuelva.com':
            send_account = account
            break

   mail_item = outlook.CreateItem(0)   # 0: olMailItem

    # mail_item.SendUsingAccount = send_account not working
    # the following statement performs the function instead
   mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))

   mail_item.Recipients.Add('alfonso.raggio@camarahuelva.com')
   mail_item.Subject = 'Test sending using particular account'
   mail_item.BodyFormat = 2   # 2: Html format 
   for file in adjuntos:      
      mail_item.Attachments.Add(file)
   for file in listdiag:
      mail_item.Attachments.Add(file)
   filename = "innocamaras.png"                   
   attachment3 =  'C:\\Users\\alfonso.DESKTOP-0B71OCJ\\OneDrive\\Desarrollo\\python\\automatizacion\\generar-anexos\\img\\innocamaras.png'
    

   attach = mail_item.Attachments.Add(attachment3)  
    

    # data:image/png;base64,

   with open(attachment3, "rb") as image:
     image_string = base64.b64encode(image.read())
            # imagede_string = base64.b64decode(image_string))

   mail_item.HTMLBody ='''
        <html><body>
        <p></p>
        <p>


    </p><p>Buenos
    días.</p>

    <p>Adjunto
    les remito la documentación a firmar para que puedan comenzar la implantación
    de los proyectos que nos indicó.</p>

    <p>Ruego comprueben bien los datos antes de su firma ya que en
    ella se estipulan los proyectos y cuantías económicas autorizadas por esta
    Cámara. <b><u>Una vez firmado por ambas partes, podrán comenzar a ejecutar
    gastos. Ningún gasto realizado con anterioridad a la firma de los documentos
    por ambas partes, no será admitido.</u></b></p>

    <p>Los
    documentos&nbsp; en formato PDF pueden remitirlos firmados con <b><u>su
    certificado digital. </u></b></p>

    <p>A continuación recibirá un email indicando las obligaciones de publicidad de
    la UE que deben cumplir.</p>

    <p>Les recordamos que los plazas de ejecución y justificación son los
    siguientes.</p>

    <table style="background-color:#FFFFE0;" border="1">
    <tbody><tr>
    <td valign="top">
    <p><b>FECHA FIN DE
    EJECUCIÓN Y PAGOS</b></p>
    </td>
    <td valign="top">
    <p><b>31 de agosto
    de 2023</b></p>
    </td>
    </tr>
    <tr>
    <td valign="top">
    <p><b>PLAZO LÍMITE
    DE JUSTIFICACIÓN</b></p>
    </td>
    <td valign="top">
    <p><b>15 de
    septiembre de 2023</b></p>
    </td>
    </tr>
    </tbody></table>

    <p>Le adjuntamos también el documento donde se detalla algunas consideraciones
    a tener en cuenta para el pago y justificación de las facturas. Recomendamos su
    lectura para evitar incidencias en la justificación.</p>

    <p>Saludos y gracias de antemano. </p>
    ''' +  '<p>' +   '</p>  <img src="data:image/png;base64, ' + '' + str(image_string.decode('utf-8')) + '">'


    # <p></p><p></p><p>&nbsp;</p> </body></html>    
   mail_item.Save()
   return







def crear_draft_anexo19():
    send_account = None
    ruta = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/anexo19_" + cifglobal + ".docx"
    convert(ruta, "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/anexo19.pdf")
    for account in outlook.Session.Accounts:
        if account.DisplayName == 'alfonso.raggio@camarahuelva.com':
            send_account = account
            break

    mail_item = outlook.CreateItem(0)   # 0: olMailItem

    # mail_item.SendUsingAccount = send_account not working
    # the following statement performs the function instead
    mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))

    mail_item.Recipients.Add('alfonso.raggio@camarahuelva.com')
    mail_item.Subject = 'Test sending using particular account'
    mail_item.BodyFormat = 2   # 2: Html format
    attachment = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/anexo19.pdf"
    mail_item.Attachments.Add(attachment)


    mail_item.HTMLBody = '''
        <H2>Hello, This is a test mail.</H2>
        Hello Guys. 
        '''

    mail_item.Save()
    return








def enviar_anexo19():
    print("opcion seleccionada enviar anexo 19")
        
    # mail = outlook.CreateItem(0) 
    mail =  outlook.CreateItem(0) 
    #mail =  outlook.CreateItemFromTemplate('C:\\temp\\prueba-oft.oft')
                                                

    mail.To = 'alfonso.raggio@camarahuelva.com'
    mail.Subject = ' asunto de correo '


    mail.HTMLBody =  """
  <html><body>
	<p></p>
	<p>


</p><p>Buenos
días.</p>

<p>Adjunto
les remito la documentación a firmar para que puedan comenzar la implantación
de los proyectos que nos indicó.</p>

<p>Ruego comprueben bien los datos antes de su firma ya que en
ella se estipulan los proyectos y cuantías económicas autorizadas por esta
Cámara. <b><u>Una vez firmado por ambas partes, podrán comenzar a ejecutar
gastos. Ningún gasto realizado con anterioridad a la firma de los documentos
por ambas partes, no será admitido.</u></b></p>

<p>Los
documentos&nbsp; en formato PDF pueden remitirlos firmados con <b><u>su
certificado digital. </u></b></p>

<p>A continuación recibirá un email indicando las obligaciones de publicidad de
la UE que deben cumplir.</p>

<p>Les recordamos que los plazas de ejecución y justificación son los
siguientes.</p>

<table style="background-color:#FFFFE0;" border="1">
 <tbody><tr>
  <td valign="top">
  <p><b>FECHA FIN DE
  EJECUCIÓN Y PAGOS</b></p>
  </td>
  <td valign="top">
  <p><b>31 de agosto
  de 2023</b></p>
  </td>
 </tr>
 <tr>
  <td valign="top">
  <p><b>PLAZO LÍMITE
  DE JUSTIFICACIÓN</b></p>
  </td>
  <td valign="top">
  <p><b>15 de
  septiembre de 2023</b></p>
  </td>
 </tr>
</tbody></table>

<p>Le adjuntamos también el documento donde se detalla algunas consideraciones
a tener en cuenta para el pago y justificación de las facturas. Recomendamos su
lectura para evitar incidencias en la justificación.</p>

<p>Saludos y gracias de antemano. </p>

<!-- HTML Code -->
<img src="C://Users//alfonso.DESKTOP-0B71OCJ//OneDrive//Desarrollo//python//automatizacion//generar-anexos//img//innocamaras.png">
  </figure><p></p><p></p> </body></html>

    """

    mail.BodyFormat = 2

    mail.Send()


    return
    
          
def mostrar_menu(opciones):
    print('Seleccione una opción:')
    for clave in opciones:
        # print(f' {opciones[clave][0]}')
         print(f' {clave}) {opciones[clave][0]}')
    
def leer_opcion(opciones):
    while (a := input('Opción: ')) not in opciones:
        print('Opción incorrecta, vuelva a intentarlo.')
    return a

def ejecutar_opcion(opcion, opciones):
    opciones[opcion][1]()
    
def generar_menu(opciones, opcion_salida):
    opcion = None    
    while opcion != opcion_salida:
        mostrar_menu(opciones)
        opcion = leer_opcion(opciones)
        ejecutar_opcion(opcion, opciones)
        print(" el cif en el while es " + cifglobal)
        print () # se imprime opcion en blanco para clarificar salir de pantalla

def generar_ficha_oap():   
    # cif=input()tr


    cif_a_buscar = 'B21542931'
    cif = '44216983n' 
    #context = {'mi_nombre': 'Fran tarci'}
    df_tab_LISTA_ASESORE_GENERACION_DOCUMEs = read_data_pandas_oap()    
    
    # se localiza por NIF en el dataframe que corresponde al tab ALTA EMPRESAS    
    df_tab_empresa_fila_buscada = df_tab_LISTA_ASESORE_GENERACION_DOCUMEs.loc[df_tab_LISTA_ASESORE_GENERACION_DOCUMEs['documento_soclitiante'] == cif_a_buscar]   
    
    # se localiza por NIF en el dataframe que corresponde al tab ALTA USUARIO
    # df_tab_usuario_buscado = df_tab_alta_usuarios.loc[df_tab_alta_usuarios['documento_soclitiante'] == cif_a_buscar]
    
    #print(" kkkk" + df_tab_LISTA_ASESORE_GENERACION_DOCUMEs.to_string())

    print("empresa encontrada para imprimir ")
    #print(df_tab_empresa_fila_buscada)
    print("USUARIO BUSCADO ============> usuario encontrada para imprimir ")
    # print(df_tab_usuario_buscado)
    # buscar valor en un dataframe por columna
    #
    # con at df1.at[0,'randomcolumn'] => el AT NO FUNCIONA
    # He probado con iloc y funciona
    #==========================================
    #================= DATOS DE EMPRESA ==========
    # ============================================
    cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    nombre_representante = df_tab_empresa_fila_buscada.iloc[0]['nombre_representante']
    
    #empleados_empresa = df_tab_empresa_fila_buscada.iloc[0]['Empleados empresa']
    

    # ====================================
    # ==============   datos ALTA USAURIO 
    # ====================================
    # Razón social  - calculado
    # nifempresaE  = df_tab_usuario_buscado.iloc[0]['documento_soclitiante']       
      # Estado NIF - calculado
    #tratamiento =   df_tab_usuario_buscado.iloc[0]['Tratamiento']
   
   
     

    ##### cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    ##### nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']     
    print('cif de empresa ' + str(cif_empresa))
    print('nombre de empresa  ' + nombre_o_razon_social) 

    context = {
                
                 # =============== FICHA DE EMPRESA
                'nif_de_empresa' : cif_empresa,
                'razon_social' : nombre_o_razon_social       , #revisar si intruso
                'nombre_representante' : nombre_representante
                # empleados que es un ccheckbox
                # 'inicio_de actividad'  son cuadros
                # sector actividad   son cuadros
                #'domicilio_social' :  'pendiente',
                #'codigo_postal' : codigopostal,
                #'localidad' : localidad,
                #'provincia' : provincia,
                #'email_general'  : correoelectronico,
                # 'pagina web' : paginaweb,
                # ================= DATOS DE CONSULTA
                #'nombre' : nombre,
                #'Primer_apellido' : primerapellido,
                #'Segundo_apellido' : segundoapellido,
                #'fechaasesoramiento' : str(fechaasesoramiento),
                #'dni' : nif,
                #'cargo' : cargo,
                #'email_contacto' : correoelectronico,
                #'fecha_consulta' : fechaasesoramiento,
                #'como_nos_ha_conocido' : comonoshaconocido,
                #'num_assistentes' : 'n/a',
                #'canalservicio' : correoelectronico,
                #'tematica' : 'tematica a tratar',
                #'observaciones' : 'Incluir observaciones de la consulta 
                
                
    }

   

    path = str(os.getcwd())
    # si quisieramos cambiar la ruta os.getcwd(path)
    print("la ruta de trabajo es " + path)
    doc = DocxTemplate('Anexo 03. Ficha de Empresa 2022-v3.docx')        
    doc.render(context)
    doc.save('generated_doc'+ cif_empresa + '.docx')
      

            
print("=============== cuerpo principal ========================")


def valuestovar(df):
    df_tab_empresa_fila_buscada = df
    programa = df_tab_empresa_fila_buscada.iloc[0]['programa']
    fases = df_tab_empresa_fila_buscada.iloc[0]['Fases']
    nombre_solicitante = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    provincia = df_tab_empresa_fila_buscada.iloc[0]['provincia']
    poblacion = df_tab_empresa_fila_buscada.iloc[0]['poblacion']
    cp = df_tab_empresa_fila_buscada.iloc[0]['cp']
    email = df_tab_empresa_fila_buscada.iloc[0]['email']
    direccion = df_tab_empresa_fila_buscada.iloc[0]['direccion']
    telefono_solicitante = df_tab_empresa_fila_buscada.iloc[0]['telefono_solicitante']
    documento_representante = df_tab_empresa_fila_buscada.iloc[0]['docoumento_representante']
    email_representante = df_tab_empresa_fila_buscada.iloc[0]['email_representante']
    cargo = df_tab_empresa_fila_buscada.iloc[0]['cargo']
    tratamiento_representante = df_tab_empresa_fila_buscada.iloc[0]['tratamiento_representante']
    fases2 = df_tab_empresa_fila_buscada.iloc[0]['fases2']
    tecnico_justificar = df_tab_empresa_fila_buscada.iloc[0]['tecnico_justificar']
    dni_tecnico = df_tab_empresa_fila_buscada.iloc[0]['dni_tecnico']
    fecha_documento_inicio_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_documento_inicio_diagnostico']
    fecha_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_diagnostico']
    anexo_18_firmado = df_tab_empresa_fila_buscada.iloc[0]['anexo_18_firmado']
    fecha_recepcion_presupuesto = df_tab_empresa_fila_buscada.iloc[0]['fecha_recepcion_presupuesto']
    fecha_envio_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_envio_anexo_19']
    envio_ppi = df_tab_empresa_fila_buscada.iloc[0]['envio_ppi']
    fecha_firma_ppi = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_ppi']
    fecha_firma_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_anexo_19']
    duracion_del_plan = df_tab_empresa_fila_buscada.iloc[0]['duracion_del_plan']
    encuesta_satisfaccion = df_tab_empresa_fila_buscada.iloc[0]['encuesta_satisfaccion']
    proyectos = df_tab_empresa_fila_buscada.iloc[0]['proyectos']
    descripcion_empresa = df_tab_empresa_fila_buscada.iloc[0]['descripcion_empresa']
    fecha_registro_participacion_fase_i_anexo_18 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_fase_i_anexo_18']
    enviado_email_publicidad_ue = df_tab_empresa_fila_buscada.iloc[0]['enviado_email_publicidad_ue']
    fecha_registro_participacion_en_fase_ii_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_en_fase_ii_anexo_19']    
    cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    nombre_representante = df_tab_empresa_fila_buscada.iloc[0]['nombre_representante']
    ruta_en_disco = df_tab_empresa_fila_buscada.iloc[0]['ruta_en_disco']
     

    
    #empleados_empresa = df_tab_empresa_fila_buscada.iloc[0]['Empleados empresa']
    

    # ====================================
    # ==============   datos ALTA USAURIO 
    # ====================================
    # Razón social  - calculado
    # nifempresaE  = df_tab_usuario_buscado.iloc[0]['documento_soclitiante']       
      # Estado NIF - calculado
    #tratamiento =   df_tab_usuario_buscado.iloc[0]['Tratamiento']
   
   
     

    ##### cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    ##### nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']     
    print('cif de empresa ' + str(cif_empresa))
    print('nombre de empresa  ' + nombre_o_razon_social) 
    
    print ('Fecha del documento --> ' + str(fecha_documento))
    print ( ' REPRESENTANTE ' + str(nombre_representante.lower()))

    context = {
                'fecha_documento' : fecha_documento,
                 # =============== FICHA DE EMPRESA
                'nif_de_empresa' : cif_empresa,
                'nombre_representante' : nombre_representante.lower().title(),
                'razon_social' : nombre_o_razon_social       , #revisar si intruso
                'tecnico_justificar' :  tecnico_justificar,

                'programa' : programa,
                'fases' : fases,
                'nombre_solicitante' : nombre_solicitante,
                'provincia' : provincia,
                'poblacion' : poblacion,
                'cp' : cp,
                'email' : email,
                'direccion' : direccion,
                'telefono_solicitante' : telefono_solicitante,
                'documento_representante' : documento_representante,
                'email_representante' : email_representante,
                'cargo' : cargo,
                'tratamiento_representante' : tratamiento_representante,
                'fases2' : fases2,
                'tecnico_justificar' : tecnico_justificar,
                'dni_tecnico' : dni_tecnico,
                'fecha_documento_inicio_diagnostico' : fecha_documento_inicio_diagnostico,
                'fecha_diagnostico' : fecha_diagnostico,
                'anexo_18_firmado' : anexo_18_firmado,
                'fecha_recepcion_presupuesto' : fecha_recepcion_presupuesto,
                'fecha_envio_anexo_19' : fecha_envio_anexo_19,
                'envio_ppi' : envio_ppi,
                'fecha_firma_ppi' : fecha_firma_ppi,
                'fecha_firma_anexo_19' : fecha_firma_anexo_19,
                'duracion_del_plan' : duracion_del_plan,
                'encuesta_satisfaccion' : encuesta_satisfaccion,
                'proyectos' : proyectos,
                'descripcion_empresa' : descripcion_empresa,
                'fecha_registro_participacion_fase_i_anexo_18' : fecha_registro_participacion_fase_i_anexo_18,
                'enviado_email_publicidad_ue' : enviado_email_publicidad_ue,
                'fecha_registro_participacion_en_fase_ii_anexo_19':   fecha_registro_participacion_en_fase_ii_anexo_19,
                'cif_empresa' :cif_empresa,
                'nombre_o_razon_social' :  nombre_o_razon_social,
                

    






                # empleados que es un ccheckbox
                # 'inicio_de actividad'  son cuadros
                # sector actividad   son cuadros
                #'domicilio_social' :  'pendiente',
                #'codigo_postal' : codigopostal,
                #'localidad' : localidad,
                #'provincia' : provincia,
                #'email_general'  : correoelectronico,
                # 'pagina web' : paginaweb,
                # ================= DATOS DE CONSULTA
                #'nombre' : nombre,
                #'Primer_apellido' : primerapellido,
                #'Segundo_apellido' : segundoapellido,
                #'fechaasesoramiento' : str(fechaasesoramiento),
                #'dni' : nif,
                #'cargo' : cargo,
                #'email_contacto' : correoelectronico,
                #'fecha_consulta' : fechaasesoramiento,
                #'como_nos_ha_conocido' : comonoshaconocido,
                #'num_assistentes' : 'n/a',
                #'canalservicio' : correoelectronico,
                #'tematica' : 'tematica a tratar',
                #'observaciones' : 'Incluir observaciones de la consulta 
                
                
    }

   

    path = str(os.getcwd())
    # si quisieramos cambiar la ruta os.getcwd(path)
    print("la ruta de trabajo es " + path)
    doc = DocxTemplate('anexo19.docx')        
    doc.render(context)
    doc.save('generated_anexo19_'+ cifglobal + '.docx')           
    return



def generar_ficha_anexo19():   
    # cif=input()
    cif_a_buscar = cifglobal
 
    #context = {'mi_nombre': 'Fran tarci'}
    df_tab_LISTA_ASESORE_GENERACION_DOCUMEs = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME   
    print (('en generar ficha '+ str(df_tab_LISTA_ASESORE_GENERACION_DOCUMEs)))

    
    # se localiza por NIF en el dataframe que corresponde al tab ALTA EMPRESAS    
    df_tab_empresa_fila_buscada = df_tab_LISTA_ASESORE_GENERACION_DOCUMEs.loc[df_tab_LISTA_ASESORE_GENERACION_DOCUMEs['documento_soclitiante'] == cif_a_buscar]   

    valuestovar(df_tab_empresa_fila_buscada)
    """  
    # se localiza por NIF en el dataframe que corresponde al tab ALTA USUARIO
    # df_tab_usuario_buscado = df_tab_alta_usuarios.loc[df_tab_alta_usuarios['documento_soclitiante'] == cif_a_buscar]
    
    #print(" kkkk" + df_tab_LISTA_ASESORE_GENERACION_DOCUMEs.to_string())

    print("empresa encontrada para imprimir ")
    #print(df_tab_empresa_fila_buscada)
    print("USUARIO BUSCADO ============> usuario encontrada para imprimir ")
    # print(df_tab_usuario_buscado)
    # buscar valor en un dataframe por columna
    #
    # con at df1.at[0,'randomcolumn'] => el AT NO FUNCIONA
    # He probado con iloc y funciona
    #==========================================
    #================= DATOS DE EMPRESA ==========
    # ============================================
    programa = df_tab_empresa_fila_buscada.iloc[0]['programa']
    fases = df_tab_empresa_fila_buscada.iloc[0]['Fases']
    nombre_solicitante = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    provincia = df_tab_empresa_fila_buscada.iloc[0]['provincia']
    poblacion = df_tab_empresa_fila_buscada.iloc[0]['poblacion']
    cp = df_tab_empresa_fila_buscada.iloc[0]['cp']
    email = df_tab_empresa_fila_buscada.iloc[0]['email']
    direccion = df_tab_empresa_fila_buscada.iloc[0]['direccion']
    telefono_solicitante = df_tab_empresa_fila_buscada.iloc[0]['telefono_solicitante']
    documento_representante = df_tab_empresa_fila_buscada.iloc[0]['docoumento_representante']
    email_representante = df_tab_empresa_fila_buscada.iloc[0]['email_representante']
    cargo = df_tab_empresa_fila_buscada.iloc[0]['cargo']
    tratamiento_representante = df_tab_empresa_fila_buscada.iloc[0]['tratamiento_representante']
    fases2 = df_tab_empresa_fila_buscada.iloc[0]['fases2']
    tecnico_justificar = df_tab_empresa_fila_buscada.iloc[0]['tecnico_justificar']
    dni_tecnico = df_tab_empresa_fila_buscada.iloc[0]['dni_tecnico']
    fecha_documento_inicio_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_documento_inicio_diagnostico']
    fecha_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_diagnostico']
    anexo_18_firmado = df_tab_empresa_fila_buscada.iloc[0]['anexo_18_firmado']
    fecha_recepcion_presupuesto = df_tab_empresa_fila_buscada.iloc[0]['fecha_recepcion_presupuesto']
    fecha_envio_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_envio_anexo_19']
    envio_ppi = df_tab_empresa_fila_buscada.iloc[0]['envio_ppi']
    fecha_firma_ppi = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_ppi']
    fecha_firma_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_anexo_19']
    duracion_del_plan = df_tab_empresa_fila_buscada.iloc[0]['duracion_del_plan']
    encuesta_satisfaccion = df_tab_empresa_fila_buscada.iloc[0]['encuesta_satisfaccion']
    proyectos = df_tab_empresa_fila_buscada.iloc[0]['proyectos']
    descripcion_empresa = df_tab_empresa_fila_buscada.iloc[0]['descripcion_empresa']
    fecha_registro_participacion_fase_i_anexo_18 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_fase_i_anexo_18']
    enviado_email_publicidad_ue = df_tab_empresa_fila_buscada.iloc[0]['enviado_email_publicidad_ue']
    fecha_registro_participacion_en_fase_ii_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_en_fase_ii_anexo_19']    
    cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    nombre_representante = df_tab_empresa_fila_buscada.iloc[0]['nombre_representante']

    
    #empleados_empresa = df_tab_empresa_fila_buscada.iloc[0]['Empleados empresa']
    

    # ====================================
    # ==============   datos ALTA USAURIO 
    # ====================================
    # Razón social  - calculado
    # nifempresaE  = df_tab_usuario_buscado.iloc[0]['documento_soclitiante']       
      # Estado NIF - calculado
    #tratamiento =   df_tab_usuario_buscado.iloc[0]['Tratamiento']
   
   
     

    ##### cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    ##### nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']     
    print('cif de empresa ' + str(cif_empresa))
    print('nombre de empresa  ' + nombre_o_razon_social) 
    
    print ('Fecha del documento --> ' + str(fecha_documento))
    print ( ' REPRESENTANTE ' + str(nombre_representante.lower()))

    context = {
                'fecha_documento' : fecha_documento,
                 # =============== FICHA DE EMPRESA
                'nif_de_empresa' : cif_empresa,
                'nombre_representante' : nombre_representante.lower().title(),
                'razon_social' : nombre_o_razon_social       , #revisar si intruso
                'tecnico_justificar' :  tecnico_justificar,

                'programa' : programa,
                'fases' : fases,
                'nombre_solicitante' : nombre_solicitante,
                'provincia' : provincia,
                'poblacion' : poblacion,
                'cp' : cp,
                'email' : email,
                'direccion' : direccion,
                'telefono_solicitante' : telefono_solicitante,
                'documento_representante' : documento_representante,
                'email_representante' : email_representante,
                'cargo' : cargo,
                'tratamiento_representante' : tratamiento_representante,
                'fases2' : fases2,
                'tecnico_justificar' : tecnico_justificar,
                'dni_tecnico' : dni_tecnico,
                'fecha_documento_inicio_diagnostico' : fecha_documento_inicio_diagnostico,
                'fecha_diagnostico' : fecha_diagnostico,
                'anexo_18_firmado' : anexo_18_firmado,
                'fecha_recepcion_presupuesto' : fecha_recepcion_presupuesto,
                'fecha_envio_anexo_19' : fecha_envio_anexo_19,
                'envio_ppi' : envio_ppi,
                'fecha_firma_ppi' : fecha_firma_ppi,
                'fecha_firma_anexo_19' : fecha_firma_anexo_19,
                'duracion_del_plan' : duracion_del_plan,
                'encuesta_satisfaccion' : encuesta_satisfaccion,
                'proyectos' : proyectos,
                'descripcion_empresa' : descripcion_empresa,
                'fecha_registro_participacion_fase_i_anexo_18' : fecha_registro_participacion_fase_i_anexo_18,
                'enviado_email_publicidad_ue' : enviado_email_publicidad_ue,
                'fecha_registro_participacion_en_fase_ii_anexo_19':   fecha_registro_participacion_en_fase_ii_anexo_19,
                'cif_empresa' :cif_empresa,
                'nombre_o_razon_social' :  nombre_o_razon_social,
                

    






                # empleados que es un ccheckbox
                # 'inicio_de actividad'  son cuadros
                # sector actividad   son cuadros
                #'domicilio_social' :  'pendiente',
                #'codigo_postal' : codigopostal,
                #'localidad' : localidad,
                #'provincia' : provincia,
                #'email_general'  : correoelectronico,
                # 'pagina web' : paginaweb,
                # ================= DATOS DE CONSULTA
                #'nombre' : nombre,
                #'Primer_apellido' : primerapellido,
                #'Segundo_apellido' : segundoapellido,
                #'fechaasesoramiento' : str(fechaasesoramiento),
                #'dni' : nif,
                #'cargo' : cargo,
                #'email_contacto' : correoelectronico,
                #'fecha_consulta' : fechaasesoramiento,
                #'como_nos_ha_conocido' : comonoshaconocido,
                #'num_assistentes' : 'n/a',
                #'canalservicio' : correoelectronico,
                #'tematica' : 'tematica a tratar',
                #'observaciones' : 'Incluir observaciones de la consulta 
                
                
    }

   

    path = str(os.getcwd())
    # si quisieramos cambiar la ruta os.getcwd(path)
    print("la ruta de trabajo es " + path)
    doc = DocxTemplate('anexo19.docx')        
    doc.render(context)
    doc.save('generated_anexo19_'+ cifglobal + '.docx')           
    return
 """

def procbuscarcif():
    print('Introduzca CIF ')
    
    # cif=input()
    #cif = '44237153B'
    global cifglobal
    cifglobal = input(" Introducir CIF ")
    
    # ========= Si lo encontramos, creamos un dataframe con esa fila para la pestaña EMPRESA y USUARIO  
    # == Cargamos en DF datos TAB EMPRESA
    print('')
    #df_empresa_buscada = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME.loc[data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME['documento_soclitiante'] == cif]      
    
    print('El CIF es ' + cif)
    # dfAltaEmpresas, dfAltaUsuario = modify_data_pandas(cif)
    #,data_oap_tab_alta_usuario)
    #update_spreadsheet(path, CF, 1, 1, 'Alta_Usuario') #Write to sheet1 starting from row 20 and column 3 / column C
    cif_a_buscar = cifglobal
 
    #context = {'mi_nombre': 'Fran tarci'}
    df_tab_LISTA_ASESORE_GENERACION_DOCUMEs = data_oap_tab_LISTA_ASESORE_GENERACION_DOCUME   
    print (('en generar ficha '+ str(df_tab_LISTA_ASESORE_GENERACION_DOCUMEs)))

    
    # se localiza por NIF en el dataframe que corresponde al tab ALTA EMPRESAS    
    df_tab_empresa_fila_buscada = df_tab_LISTA_ASESORE_GENERACION_DOCUMEs.loc[df_tab_LISTA_ASESORE_GENERACION_DOCUMEs['documento_soclitiante'] == cif_a_buscar]   

    #valuestovar(df_tab_empresa_fila_buscada)


  
    programa = df_tab_empresa_fila_buscada.iloc[0]['programa']
    fases = df_tab_empresa_fila_buscada.iloc[0]['Fases']
    nombre_solicitante = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    provincia = df_tab_empresa_fila_buscada.iloc[0]['provincia']
    poblacion = df_tab_empresa_fila_buscada.iloc[0]['poblacion']
    cp = df_tab_empresa_fila_buscada.iloc[0]['cp']
    email = df_tab_empresa_fila_buscada.iloc[0]['email']
    direccion = df_tab_empresa_fila_buscada.iloc[0]['direccion']
    telefono_solicitante = df_tab_empresa_fila_buscada.iloc[0]['telefono_solicitante']
    documento_representante = df_tab_empresa_fila_buscada.iloc[0]['docoumento_representante']
    email_representante = df_tab_empresa_fila_buscada.iloc[0]['email_representante']
    cargo = df_tab_empresa_fila_buscada.iloc[0]['cargo']
    tratamiento_representante = df_tab_empresa_fila_buscada.iloc[0]['tratamiento_representante']
    fases2 = df_tab_empresa_fila_buscada.iloc[0]['fases2']
    tecnico_justificar = df_tab_empresa_fila_buscada.iloc[0]['tecnico_justificar']
    dni_tecnico = df_tab_empresa_fila_buscada.iloc[0]['dni_tecnico']
    fecha_documento_inicio_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_documento_inicio_diagnostico']
    fecha_diagnostico = df_tab_empresa_fila_buscada.iloc[0]['fecha_diagnostico']
    anexo_18_firmado = df_tab_empresa_fila_buscada.iloc[0]['anexo_18_firmado']
    fecha_recepcion_presupuesto = df_tab_empresa_fila_buscada.iloc[0]['fecha_recepcion_presupuesto']
    fecha_envio_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_envio_anexo_19']
    envio_ppi = df_tab_empresa_fila_buscada.iloc[0]['envio_ppi']
    fecha_firma_ppi = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_ppi']
    fecha_firma_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_firma_anexo_19']
    duracion_del_plan = df_tab_empresa_fila_buscada.iloc[0]['duracion_del_plan']
    encuesta_satisfaccion = df_tab_empresa_fila_buscada.iloc[0]['encuesta_satisfaccion']
    proyectos = df_tab_empresa_fila_buscada.iloc[0]['proyectos']
    descripcion_empresa = df_tab_empresa_fila_buscada.iloc[0]['descripcion_empresa']
    fecha_registro_participacion_fase_i_anexo_18 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_fase_i_anexo_18']
    enviado_email_publicidad_ue = df_tab_empresa_fila_buscada.iloc[0]['enviado_email_publicidad_ue']
    fecha_registro_participacion_en_fase_ii_anexo_19 = df_tab_empresa_fila_buscada.iloc[0]['fecha_registro_participacion_en_fase_ii_anexo_19']    
    cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']
    nombre_representante = df_tab_empresa_fila_buscada.iloc[0]['nombre_representante']
    ruta_en_disco = df_tab_empresa_fila_buscada.iloc[0]['ruta_en_disco']
    


    global rutaendiscoglobal
    rutaendiscoglobal = ruta_en_disco
    print ( "ruta en disco global dentro " + rutaendiscoglobal )

    
    #empleados_empresa = df_tab_empresa_fila_buscada.iloc[0]['Empleados empresa']
    

    # ====================================
    # ==============   datos ALTA USAURIO 
    # ====================================
    # Razón social  - calculado
    # nifempresaE  = df_tab_usuario_buscado.iloc[0]['documento_soclitiante']       
      # Estado NIF - calculado
    #tratamiento =   df_tab_usuario_buscado.iloc[0]['Tratamiento']
   
   
     

    ##### cif_empresa = df_tab_empresa_fila_buscada.iloc[0]['documento_soclitiante']
    ##### nombre_o_razon_social = df_tab_empresa_fila_buscada.iloc[0]['nombre_solicitante']     
    print('cif de empresa ' + str(cif_empresa))
    print('nombre de empresa  ' + nombre_o_razon_social) 
    
    print ('Fecha del documento --> ' + str(fecha_documento))
    print ( ' REPRESENTANTE ' + str(nombre_representante.lower()))

    context = {
                'fecha_documento' : fecha_documento,
                 # =============== FICHA DE EMPRESA
                'nif_de_empresa' : cif_empresa,
                'nombre_representante' : nombre_representante.lower().title(),
                'razon_social' : nombre_o_razon_social       , #revisar si intruso
                'tecnico_justificar' :  tecnico_justificar,

                'programa' : programa,
                'fases' : fases,
                'nombre_solicitante' : nombre_solicitante,
                'provincia' : provincia,
                'poblacion' : poblacion,
                'cp' : cp,
                'email' : email,
                'direccion' : direccion,
                'telefono_solicitante' : telefono_solicitante,
                'documento_representante' : documento_representante,
                'email_representante' : email_representante,
                'cargo' : cargo,
                'tratamiento_representante' : tratamiento_representante,
                'fases2' : fases2,
                'tecnico_justificar' : tecnico_justificar,
                'dni_tecnico' : dni_tecnico,
                'fecha_documento_inicio_diagnostico' : fecha_documento_inicio_diagnostico,
                'fecha_diagnostico' : fecha_diagnostico,
                'anexo_18_firmado' : anexo_18_firmado,
                'fecha_recepcion_presupuesto' : fecha_recepcion_presupuesto,
                'fecha_envio_anexo_19' : fecha_envio_anexo_19,
                'envio_ppi' : envio_ppi,
                'fecha_firma_ppi' : fecha_firma_ppi,
                'fecha_firma_anexo_19' : fecha_firma_anexo_19,
                'duracion_del_plan' : duracion_del_plan,
                'encuesta_satisfaccion' : encuesta_satisfaccion,
                'proyectos' : proyectos,
                'descripcion_empresa' : descripcion_empresa,
                'fecha_registro_participacion_fase_i_anexo_18' : fecha_registro_participacion_fase_i_anexo_18,
                'enviado_email_publicidad_ue' : enviado_email_publicidad_ue,
                'fecha_registro_participacion_en_fase_ii_anexo_19':   fecha_registro_participacion_en_fase_ii_anexo_19,
                'cif_empresa' :cif_empresa,
                'nombre_o_razon_social' :  nombre_o_razon_social,
                'ruta_en_disco' :  ruta_en_disco
                

    






                # empleados que es un ccheckbox
                # 'inicio_de actividad'  son cuadros
                # sector actividad   son cuadros
                #'domicilio_social' :  'pendiente',
                #'codigo_postal' : codigopostal,
                #'localidad' : localidad,
                #'provincia' : provincia,
                #'email_general'  : correoelectronico,
                # 'pagina web' : paginaweb,
                # ================= DATOS DE CONSULTA
                #'nombre' : nombre,
                #'Primer_apellido' : primerapellido,
                #'Segundo_apellido' : segundoapellido,
                #'fechaasesoramiento' : str(fechaasesoramiento),
                #'dni' : nif,
                #'cargo' : cargo,
                #'email_contacto' : correoelectronico,
                #'fecha_consulta' : fechaasesoramiento,
                #'como_nos_ha_conocido' : comonoshaconocido,
                #'num_assistentes' : 'n/a',
                #'canalservicio' : correoelectronico,
                #'tematica' : 'tematica a tratar',
                #'observaciones' : 'Incluir observaciones de la consulta 
                
                
    }

   



    return






# no usado #####################################
# no usado =======================================
# no usado #####################################

# ==============================================================
# ==============================================================
# ==============================================================


def fechadocumento():
    print('Fecha documento a generar ')
    
    # cif=input()
    #cif = '44237153B'
    global fecha_documento
    fecha_documento = input(" Introducir fecha documento ")
    
def listar_ddp():
   rutaorigen  = 'C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/'
   global rutaendiscoglobal
   print ( "ruta en disco global fuera " + rutaendiscoglobal ) 
   listddp = glob.glob('*ddp*.docx')
   listdiag = glob.glob('diagnostico*.docx')
   hayddp = "yes"
   haydiag = "yes"
   if listddp: 
     print(listddp)
   else:
        hayddp = "no"
        user_input = input(" No hay ddp para mover, seguir ?")
        if user_input.lower() == 'yes':
            print('user typed yes')           
        elif user_input.lower() == 'no':
            return
        listddp.append("no ddp")
   if listdiag:
     print(listdiag[0])
   else: 
     haydiag = "no"
     user_input = input(" No hay DIAG para mover, seguir ?")
     listdiag.append("no diag")
   
     if user_input.lower() == 'yes':
            print('user typed yes')           
     elif user_input.lower() == 'no':
            return

   print(rutaendiscoglobal)
   wait = input("Press Enter to continue." )
   print("something")
   if haydiag == "yes":
        shutil.move(rutaorigen + listdiag[0], rutaendiscoglobal + '/' + listdiag[0])
   if hayddp == "yes":    
        for file in listddp:
            print(file)
            print("directorio origen"  + str(rutaorigen) + str(file) )
            shutil.move(rutaorigen + file,  rutaendiscoglobal + '/' +  file)

def preparar_escenario_ddp(): 
   rutaorigen  = 'C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos'
   ruta_ddp_diag_test  = r'C:\Users\alfonso.DESKTOP-0B71OCJ\OneDrive\Desarrollo\python\automatizacion\generar-anexos\ddp-diag-test'
   global rutaendiscoglobal
   buscar = ruta_ddp_diag_test +  '\*ddp*.docx'
   listddp = glob.glob( buscar )
   print(listddp)
   wait = input("Press Enter to continue." )
   
   listdiag = glob.glob(ruta_ddp_diag_test +  '/diagnostico*.docx')
   for file in listddp:
        print("fichero a MOVER " + file)
        wait = input("Press Enter to continue " )
        print("something")
        print("directorio origen"  + str(rutaorigen) + str(file) )
        shutil.copy(file,  rutaorigen + '/' +  os.path.basename(file))
   for file in listdiag:
        print(file)
        print("directorio origen"  + str(rutaorigen) + str(file) )
        shutil.copy(file,  rutaorigen + '/' +  os.path.basename(file))


  





def recordatorio_inno():
    send_account = None
    
    for account in outlook.Session.Accounts:
        if account.DisplayName == 'alfonso.raggio@camarahuelva.com':
            send_account = account
            break

    mail_item = outlook.CreateItem(0)   # 0: olMailItem

    # mail_item.SendUsingAccount = send_account not working
    # the following statement performs the function instead
    mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))

    mail_item.Recipients.Add('alfonso.raggio@camarahuelva.com')
    mail_item.Subject = 'IMPORTATNTE - RECORDATORIO - Programa InnoCámaras 2023 - Plazos de ejecución y tramites a seguir'
    mail_item.BodyFormat = 2   # 2: Html format
    attachment = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/doc-convocatoria/Justific@-Guia_de_usuario.pdf"
    attachment2 = "C:/Users/alfonso.DESKTOP-0B71OCJ/OneDrive/Desarrollo/python/automatizacion/generar-anexos/doc-convocatoria/05-Anexo IV. Condiciones de participación, justificación y gastos elegibles Fase de Ayudas.pdf"
    mail_item.Attachments.Add(attachment)
    mail_item.Attachments.Add(attachment2)
    filename = "innocamaras.png"                   
    attachment3 =  'C:\\Users\\alfonso.DESKTOP-0B71OCJ\\OneDrive\\Desarrollo\\python\\automatizacion\\generar-anexos\\img\\innocamaras.png'
    

    attach = mail_item.Attachments.Add(attachment3)  
    attach.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "MyId1")

    # data:image/png;base64,

    with open(attachment3, "rb") as image:
            image_string = base64.b64encode(image.read())
            # imagede_string = base64.b64decode(image_string))

    


    mail_item.HTMLBody ='''
<html><body>


<p>Nos encontramos próximos a la recta final de la implantación de proyectos y
sois muy pocas las empresas que habéis comenzado a implantar las soluciones
elegidas. Por este motivo, os recordamos los plazos de ejecución y
justificación, así como algunas consideraciones de interés.</p>

<ol type="1">
 <li>En primer lugar os recordamos los pasos a seguir una
     vez hayáis finalizado la fase de diagnóstico Fase I y entregado firmado el
     Anexo 18 - Acta recepción y cuestionario de satisfacción Fase I.</li>
 <ol type="1">
  <li>Entregar los
      presupuestos de los proyectos que queráis poner en marcha.</li>
  <li>Firmar la documentación
      de inicio de la fase de implantación que os proporcionará Cámara de
      Comercio una vez recibidos y dado el visto bueno a los presupuestos.</li>
  <li>Hasta que no se firme
      esta documentación no se podrá comenzar a ejecutar gastos del proyecto.</li>
 </ol>
</ol>

<ol type="1">
 <li>Los plazos a cumplir son los siguientes:</li>
 <ol type="1">
  <li><b>15 de agosto de 2023</b> - Fecha tope para
      enviar a la Cámara de Comercio las facturas de los gastos a justificar.
      Serán enviadas por email a la Cámara de Comercio de Huelva a través de la
      dirección <a href="mailto:innocamaras@camarahuelva.com">innocamaras@camarahuelva.com</a></li>
  <li><b>31 de agosto de 2023</b> - Fecha tope para
      realizar los pagos de todos los gastos que se vayan a presentar en la
      justificación. Se recomienda la consulta del&nbsp; anexo IV a la
      convocatoria.</li>
  <li>Una vez recibidas las
      facturas, la Cámara le remitirá los anexos 20 (Registro Prestación
      Servicio y Seguimiento - Fase II) y 21 (Memoria Ejecución Proyecto y
      Cuestionarios) que deberán ser cumplimentados y firmados para ser
      remitidos a la Cámara antes del <b>15 de septiembre de 2023 </b>a través
      de la dirección <a href="mailto:innocamaras@camarahuelva.com">innocamaras@camarahuelva.com</a></li>
  <li>Toda la documentación
      para la justificación deberá ser aportada antes del <b>15 de septiembre
      de 2023</b>. A partir de esta fecha no se podrá aportar documentación salvo
      en el caso de que se le comunique la necesidad de subsanar alguna de la
      documentación ya aportada. Esta documentación debe ser subida a la
      plataforma justific@.</li>
  <li>La relación de
      documentación a aportar a través de la plataforma justific@ es la
      siguiente:</li>
  <ol type="1">
   <li>Facturas.</li>
   <li>Justificantes de pago.
       Orden de transferencia + Extracto bancario. Puede ser sustituido por
       certificado bancario de la transferencia realizada.</li>
   <li>Evidencias de los
       gastos realizados.</li>
   <li>Evidencias del
       cumplimiento de la publicidad del FEDER</li>
   <li>Documento de
       identificación financiera. Certificado de titularidad de la cuenta
       bancaria de abono de la ayuda.</li>
  </ol>
  <li>Además de la
      documentación que deben subir a la plataforma justific@, deberán aportar
      los anexos 20 (Registro Prestación Servicio y Seguimiento - Fase II) y 21
      (Memoria Ejecución Proyecto y Cuestionarios) debidamente cumplimentados y
      firmados, así como otros anexos relativos al asesoramiento recibido. Como
      se indicaba con anterioridad, estos anexos serán proporcionados por la
      Cámara de Comercio una vez recibamos las facturas.<b><i> </i></b></li>
 </ol>
</ol>

<p>Para poder explicaros todo el proceso de justificación, tuvo lugar una 
jornada en la que se explicaba todo el proceso. En realidad la herramienta es muy fácil y el proceso es sencillo, además siempre contareis con el apoyo
de los técnicos que os están gestionando vuestro expediente.</p>

El vídeo fue una charla &nbsp;de aproximadamente 45 minutos. &nbsp;Servirá para
entender lo más importante a la hora de subir los documentos de justificación.</u></p>

<p>Enlace para vídeo</p>

<p><a href="https://us06web.zoom.us/meeting/register/tZIkd-mhrDwrE9CMR1FDfXZpLsABjvvIsP8h">https://us06web.zoom.us/meeting/register/tZIkd-mhrDwrE9CMR1FDfXZpLsABjvvIsP8h</a></p>

<p>Como siempre nos tienen a su disposición para resolver cualquier consulta o
duda.</p>

<p>Un cordial saludo.</p>

<p>&nbsp;</p>

&nbsp;
&nbsp;&nbsp;</body></html>

''' +  '<p>' +   '</p>  <img src="data:image/png;base64, ' + '' + str(image_string.decode('utf-8')) + '">'



# <p></p><p></p><p>&nbsp;</p> </body></html>    
    mail_item.Save()
    return



    

    
def menu_principal():
    print("menu - el alor del cif es " + cif)
    opciones = {
        '1': ('1.- Fecha Documento', fechadocumento),
        '2': ('2.- Buscar por CIF', procbuscarcif),
        '3': ('3.- Generar Ficha', generar_ficha_anexo19),
        '4': ('4.- Crear draft 19 + ppi ',crear_draft_anexo19),
        '4': ('4.- Crear draft 18 + diag ',crear_draft_anexo18),
        '5': ('5.- Enviar anexo 19', enviar_anexo19),
        '6': ('6.- Listar y mover ddp - req cif', listar_ddp),
         #'7': ('6.- Listar ddp ', enviar_diag_anexo18),
        '9': ('9.- Preparar Test ddp ', preparar_escenario_ddp),
        '10': ('10.- recordatorio TIC', recordatorio_inno),
        '14': ('14.- Salir', salir)
    }    
    generar_menu(opciones, '10')



def salir():
    
    print ('saliendo ')

""" __name__ es una especial variable en python que hará que cuando el código se ejecuta desde la línea de comando no se ejecute todo sino lo que que está dentro de la condición
 """

if __name__ == '__main__':
    menu_principal()

    
    




    



